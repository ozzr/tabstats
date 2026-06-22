import os
import tempfile
import pandas as pd
from tabstat import tabstat
from tabstat.generator import TabStatGenerator
from tabstat.config import TabStatConfig


def test_html_returns_string(df_medium):
    html = tabstat(df_medium, "age | outcome", tablefmt="html")
    assert isinstance(html, str)
    assert "<table" in html.lower()


def test_html_contains_title(df_medium):
    html = tabstat(df_medium, "age | outcome",
                   tablefmt="html", title="Test Title")
    assert "Test Title" in html


def test_excel_creates_file(df_medium):
    gen = TabStatGenerator(TabStatConfig())
    result = gen.generate(df_medium, "age + sex | outcome")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        gen.to_excel(result, path, title="Table 1")
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
    finally:
        os.unlink(path)


def test_excel_merges_title_and_footnote_rows(df_medium):
    gen = TabStatGenerator(TabStatConfig())
    result = gen.generate(df_medium, "age + sex | outcome", title="Table 1", footnote="IQR note")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        gen.to_excel(result, path, title="Table 1", footnote="IQR note")
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        merged = [str(r) for r in ws.merged_cells.ranges]
        title_range = f"A1:{get_column_letter(ws.max_column)}1"
        assert title_range in merged
        assert ws["A1"].value == "Table 1"
        assert ws.cell(ws.max_row, 1).value == "IQR note"
    finally:
        os.unlink(path)


def test_excel_merges_repeated_multiindex_headers(df_medium):
    gen = TabStatGenerator(TabStatConfig())
    result = gen.generate(df_medium, "age + sex | outcome")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        gen.to_excel(result, path, title="Table 1")
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        merged = [str(r) for r in ws.merged_cells.ranges]
        assert "B2:C2" in merged
        assert ws["B2"].value == "outcome"
    finally:
        os.unlink(path)


def test_excel_no_blank_row_after_header(df_medium):
    gen = TabStatGenerator(TabStatConfig())
    result = gen.generate(df_medium, "age + sex | outcome")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        gen.to_excel(result, path, title="Table 1")
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(4, 1).value == result.iloc[0, 0]
        assert ws.cell(4, 1).value == "age"
    finally:
        os.unlink(path)


def test_excel_inserts_pvalue_between_categorical_levels():
    import numpy as np
    gen = TabStatGenerator(TabStatConfig())
    np.random.seed(123)
    n = 120
    df = pd.DataFrame({
        "PLTDIC": np.random.choice(
            ["<50,000", ">=50,000", "Unknown", "Missing"],
            size=n,
            p=[0.35, 0.35, 0.2, 0.1],
        ),
        "outcome": np.random.choice([0, 1], size=n, p=[0.7, 0.3]),
    })
    df.loc[df.sample(frac=0.1, random_state=1).index, "PLTDIC"] = pd.NA
    result = gen.generate(df, "PLTDIC | outcome")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        gen.to_excel(result, path, title="Table 1")
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        header_row = None
        for row_idx in range(1, ws.max_row + 1):
            if ws.cell(row_idx, 1).value == "PLTDIC":
                header_row = row_idx
                break
        assert header_row is not None

        p_value_cell = ws.cell(header_row, 5).value
        test_cell = ws.cell(header_row, 6).value
        assert not p_value_cell
        assert not test_cell

        expected_p = result.loc[result["Characteristic"] == "PLTDIC", "P-value"].iat[0]
        expected_test = result.loc[result["Characteristic"] == "PLTDIC", "Test"].iat[0]

        first_cat_row = header_row + 1
        assert ws.cell(first_cat_row, 5).value == expected_p
        assert ws.cell(first_cat_row, 6).value == expected_test
        assert ws.cell(header_row, 5).value is None
        assert ws.cell(header_row, 6).value is None

        from openpyxl.utils import get_column_letter
        end_cat_row = first_cat_row
        while end_cat_row + 1 <= ws.max_row:
            next_label = ws.cell(end_cat_row + 1, 1).value
            if not isinstance(next_label, str) or not next_label.startswith("⠀"):
                break
            next_label_text = next_label.replace("⠀", "").strip()
            if next_label_text == "Missing":
                break
            end_cat_row += 1

        merged = [str(r) for r in ws.merged_cells.ranges]
        assert f"{get_column_letter(5)}{first_cat_row}:{get_column_letter(5)}{end_cat_row}" in merged
        assert f"{get_column_letter(6)}{first_cat_row}:{get_column_letter(6)}{end_cat_row}" in merged

        missing_row = end_cat_row + 1
        missing_label = ws.cell(missing_row, 1).value
        assert isinstance(missing_label, str)
        assert missing_label.replace("⠀", "").strip() == "Missing"
        assert ws.cell(missing_row, 5).value is None
        assert ws.cell(missing_row, 6).value is None
    finally:
        os.unlink(path)


def test_excel_workbook_multiple_tables(df_medium):
    gen = TabStatGenerator(TabStatConfig())
    result1 = gen.generate(df_medium, "age + sex | outcome", title="Table 1")
    result2 = gen.generate(df_medium, "creat | outcome", title="Table 2")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        gen.to_excel_workbook([
            ("Table1", result1),
            ("Table2", result2),
        ], path)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Table1", "Table2"]
        assert wb["Table1"].cell(1, 1).value == "Table 1"
        assert wb["Table2"].cell(1, 1).value == "Table 2"
    finally:
        os.unlink(path)
